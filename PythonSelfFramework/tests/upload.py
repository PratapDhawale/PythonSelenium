# from selenium.webdriver.chrome import webdriver
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

file_path = "F:/Data Analyst/Python Basic/Automation_Testing/ExcelFiles/download.xlsx"
Fruit_Name = "Apple"
driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(10)
wait = WebDriverWait(driver, 10)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

# Edit the Excel with updated value

book = openpyxl.load_workbook("Downloads\\download.xlsx")
sheet = book.active
dict = {}

print(sheet.max_row)
print(sheet.max_column)

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

# CSS - div [id='row-1'] div:nth-child(4)
# Xpath - //div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']
price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
print("Price Column ID:", price_column)

# Actual_price = driver.find_element(By.XPATH, "//div[text()='{Fruit_Name}']/parent::div/parent::div/div[@id='cell-{price_column}-undefined']").text
# print(Actual_price)

xpath = f"//div[text()='{Fruit_Name}']/parent::div/parent::div/div[@id='cell-{price_column}-undefined']"
print("Generated XPath:", xpath)
wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))
Actual_price = driver.find_element(By.XPATH, xpath).text
print("Actual Price:", Actual_price)
"""
try:
    wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
    Actual_price = driver.find_element(By.XPATH, xpath).text
    print("Actual Price:", Actual_price)
except Exception as e:
    print(f"Error locating element: {str(e)}")
    with open("page_source_debug.html", "w", encoding="utf-8") as f:
        f.write(driver.page_source)
    print("Page source saved for debugging.")       
    
"""