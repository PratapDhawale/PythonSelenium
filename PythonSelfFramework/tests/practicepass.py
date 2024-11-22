import os.path
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

# file_path = "F:/Data Analyst/Python Basic/Automation_Testing/ExcelFiles/download.xlsx"
file_path = os.path.expanduser("~/Downloads/download.xlsx")
Fruit_Name = "Apple"
driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(10)
wait = WebDriverWait(driver, 10)

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID, "downloadButton").click()
while not os.path.exists(file_path):
    time.sleep(1)

book = openpyxl.load_workbook(file_path)
sheet = book.active
dict = {}

print(f"Number of rows are : {sheet.max_row}")
print(f"Number of columns are : {sheet.max_column}")

for row in sheet.iter_rows(values_only=True):
    print(row)

cell = sheet.cell(row=3, column=4)
print(cell.value)

sheet.cell(row=3, column=4).value = "947"
print(sheet.cell(row=3, column=4).value)
# print(sheet.cell(row=2, column=2).value)

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

# for i in range(1, sheet.max_row + 1):
#     if sheet.cell(row=i, column=2).value=="Apple":
#         for j in range(4,sheet.max_column + 1):
#             dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=3, column=4).value

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=2).value == "Apple":
        price_column_name = sheet.cell(row=1, column=4).value
        price_value = sheet.cell(row=i, column=4).value
        dict[price_column_name] = price_value

print(dict)


