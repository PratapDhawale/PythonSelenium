import os.path
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


def update_excel_data(filepath, searchTerm, columnName, new_value):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == columnName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:/Users/prata/Downloads/download.xlsx"
Fruit_Name = "Apple"
newvalue = "950"
column_Name = "price"
driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(10)
wait = WebDriverWait(driver, 10)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()
time.sleep(2)

# Edit the Excel with updated value
update_excel_data(file_path, Fruit_Name, column_Name, newvalue)

# upload
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
xpath = f"//div[text()='{Fruit_Name}']/parent::div/parent::div/div[@id='cell-{price_column}-undefined']"
print(f"xpath is: {xpath}")
wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))
actual_price = driver.find_element(By.XPATH, xpath).text
assert actual_price == newvalue

print(f"The new value of Apple is : {actual_price}")

# Delete the file after testing
if os.path.exists(file_path):
    os.remove(file_path)
else:
    print(f"file {file_path} not found")

driver.quit()
