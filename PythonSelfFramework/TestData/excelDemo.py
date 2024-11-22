from datetime import datetime

import openpyxl

book = openpyxl.load_workbook("F:\\Data Analyst\\Python Basic\\Automation_Testing\\ExcelFiles\\Data.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

# sheet.cell(row=1, column=2).value= "Pratap"
# print(sheet.cell(row=1, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A4'].value)

for i in range(1, sheet.max_row + 1):  # get rows
    if sheet.cell(row=i, column=1).value == "TestCase1":

        for j in range(2, sheet.max_column + 1):  # get columns

            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            # print(sheet.cell(row=i, column=j).value)
print(Dict)
# print(Dict.keys())
# print(Dict.values())

"""
for i in range(1, sheet.max_row + 1):  # Iterate through rows
    if sheet.cell(row=i, column=1).value == "TestCase1":  # Match the test case name
        for j in range(2, sheet.max_column + 1):  # Iterate through columns
            header = sheet.cell(row=1, column=j).value  # Header (column name)
            value = sheet.cell(row=i, column=j).value  # Cell value

            # Handle datetime values
            if isinstance(value, datetime):  # Check if value is a datetime object
                Dict[header] = value.strftime("%d-%m-%Y")  # Format as YYYY-MM-DD
            else:
                Dict[header] = value

# Print the final dictionary
print("Final Dictionary:", Dict)
print("Keys:", Dict.keys())
print("Values:", Dict.values()) """



