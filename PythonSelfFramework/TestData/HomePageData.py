from datetime import datetime

import openpyxl


class HomePageData:
    test_HomePage_Data = \
        [{"FullName": "Pratap Dhawale", "email_address": "prtp5190@gmail.com", "Password": "12345",
          "DOB": "05-01-1990"},
         {"FullName": "Geet Dhawale", "email_address": "geet1417@gmail.com", "Password": "12345", "DOB": "14-11-2017"},
         {"FullName": "Pranjal D", "email_address": "prnjld@gmail.com", "Password": "123", "DOB": "10-01-2020"},
         {"FullName": "Shreyas s", "email_address": "sps@gmail.com", "Password": "321", "DOB": "21-10-2020"}]

    """  @staticmethod
    def getTestData(Test_Case_Name):
        Dict = {}
        book = openpyxl.load_workbook("F:\\Data Analyst\\Python Basic\\Automation_Testing\\ExcelFiles\\Data.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # get rows
            if sheet.cell(row=i, column=1).value == Test_Case_Name:
                for j in range(2, sheet.max_column + 1):  # get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    # print(sheet.cell(row=i, column=j).value)
        return [Dict]       """

    @staticmethod
    def getTestData(Test_Case_Name):
        test_data = []  # List to store dictionaries for each test case
        book = openpyxl.load_workbook("F:\\Data Analyst\\Python Basic\\Automation_Testing\\ExcelFiles\\Data.xlsx")
        sheet = book.active

        # Iterate over the rows in the sheet
        for i in range(1, sheet.max_row + 1):  # Get rows
            if sheet.cell(row=i, column=1).value == Test_Case_Name:  # Match the test case name
                Dict = {}

                for j in range(2, sheet.max_column + 1):  # Get columns
                    header = sheet.cell(row=1, column=j).value  # Get the header (e.g., "FullName", "DOB")
                    value = sheet.cell(row=i, column=j).value  # Get the cell value

                    # Check if the header is "DOB" and the value is a date
                    if header == "DOB" and isinstance(value, datetime):
                        # Format the date as "YYYY-MM-DD"
                        Dict[header] = value.strftime("%d-%m-%Y")
                    else:
                        Dict[header] = value

                test_data.append(Dict)

        return test_data
