import openpyxl
from datetime import datetime


def getTestData(excel_file_path):
    # Load the workbook and active sheet
    book = openpyxl.load_workbook(excel_file_path)
    sheet = book.active
    test_data = []

    # Iterate through rows to collect all test cases
    for i in range(2, sheet.max_row + 1):  # Start from 2 to skip the header row
        Dict = {}
        for j in range(1, sheet.max_column + 1):  # Iterate through all columns
            header = sheet.cell(row=1, column=j).value  # Get the header (column name)
            value = sheet.cell(row=i, column=j).value  # Get the cell value

            # Check if the header is "DOB" and the value is a date
            if header == "DOB" and isinstance(value, datetime):
                Dict[header] = value.strftime("%d-%m-%Y")  # Format as DD-MM-YYYY
            else:
                Dict[header] = value

        test_data.append(Dict)  # Append the row data as a dictionary

    return test_data
